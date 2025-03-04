import pytest
import pandas as pd
import openpyxl
from unittest.mock import patch, MagicMock
from io import BytesIO
import sys
from pathlib import Path
# Define the project root (the directory containing your project)
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.functions.workoutlogger import WorkoutLogger

@pytest.fixture
def mock_data():
    # Prepare mock data as DataFrames
    entry_data = {
        'Date': ["2025-02-18", "2025-02-19"],
        'Day': ["Tuesday", "Wednesday"],
        'Exercise_ID': [101, 103],
        'Exercise_Name': ["Squat", "Bench Press"],
        'Sets': [4, 3],
        'Reps': [10, 8],
        'Weight': [225, 185],
        'Rest Time (sec)': [90, 120],
        'Effectiveness': [5, 4],
        'Went to failure': [True, False],
        'Notes': ["Felt strong", "Struggled on last rep"]
    }
    exercises_data = {
        'exercise_ID': [101, 102, 103],
        'exercise_name': ['Squat', 'Deadlift', 'Bench Press'],
        'primary_muscle': ['Quads', 'Back', 'Chest'],
        'secondary_muscle': [['Glutes,Hamstrings'],
                             ['Hamstrings,Glutes,Forearms,Traps'],
                             ['Triceps,Front Delts']],
        'equipment': ['Barbell', 'Barbell', 'Dumbell'],
        'exercise_type': ['Compound', 'Compound', 'Isolation'],
        'effectiveness': ["", "", ""],
        'max_weight': ["", "", ""],
        'max_reps': ["", "", ""]
    }
    new_exercises_data = {
        'exercise_ID': [101, 102, 103, 104],
        'exercise_name': ['Squat', 'Deadlift', 'Bench Press', 'Cable Fly'],
        'primary_muscle': ['Quads', 'Back', 'Chest', 'Chest'],
        'secondary_muscle': [['Glutes,Hamstrings'],
                             ['Hamstrings,Glutes,Forearms,Traps'],
                             ['Triceps,Front Delts'],
                             ['']],
        'equipment': ['Barbell', 'Barbell', 'Dumbell', 'Machine'],
        'exercise_type': ['Compound', 'Compound', 'Isolation', 'Isolation'],
        'effectiveness': ["", "", "", ""],
        'max_weight': ["", "", "", ""],
        'max_reps': ["", "", "", ""]
    }
    entry_df = pd.DataFrame(entry_data)
    log_df = pd.DataFrame()  # Start with an empty log
    exercises_df = pd.DataFrame(exercises_data)
    new_exercises_df = pd.DataFrame(new_exercises_data)
    return entry_df, log_df, exercises_df, new_exercises_df

@pytest.fixture
def workout_logger(monkeypatch, tmp_path, mock_data):
    # 1) Create the file paths
    entry_file = tmp_path / "mock_entry.xlsx"
    log_file = tmp_path / "mock_log.json"
    exercises_file = tmp_path / "mock_exercises.json"
    
    # 2) Create a minimal Excel file with a sheet named 'entry_sheet'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'entry_sheet'
    wb.save(entry_file)

    # 3) Initialize the logger with these file paths
    logger = WorkoutLogger(str(entry_file), str(log_file), str(exercises_file))

    # 4) Get the mock data
    entry_df, log_df, exercises_df, new_exercises_df = mock_data

    # 5) Monkeypatch pandas.read_excel and read_json so logger.load_data() 
    #    returns your in-memory data. We do this to test the rest of the logic.
    def fake_read_excel(file, sheet_name):
        return entry_df if sheet_name == 'entry_sheet' else new_exercises_df
    monkeypatch.setattr(pd, "read_excel", fake_read_excel)

    def fake_read_json(file):
        return log_df if file == str(log_file) else exercises_df
    monkeypatch.setattr(pd, "read_json", fake_read_json)

    # 6) Prevent writing to JSON on update_log or update_exercise_database
    monkeypatch.setattr(pd.DataFrame, "to_json", lambda self, *args, **kwargs: None)

    # 7) Load data using the fake readers
    logger.load_data()

    return logger, entry_df, new_exercises_df, log_df, exercises_df

def test_load_data(workout_logger):
    """Test that data is loaded correctly."""
    logger, entry_df, new_exercises_df, log_df, exercises_df = workout_logger

    # Check that DataFrames are not empty (or empty for the log)
    assert not logger.entry_df.empty
    assert not logger.new_exercises_df.empty
    assert logger.log_df.empty
    assert not logger.exercises_df.empty

    # Verify that the loaded data matches our mock data
    assert logger.entry_df.equals(entry_df)
    assert logger.new_exercises_df.equals(new_exercises_df)
    assert logger.exercises_df.equals(exercises_df)

def test_update_log(monkeypatch, workout_logger):
    """Test that new workout entries are appended to the log."""
    logger, entry_df, _, log_df, _ = workout_logger

    initial_length = len(logger.log_df)
    initial_max_id = logger.log_df['log_ID'].max() if not logger.log_df.empty else 0

    # Update the log with the new entries
    logger.update_log()

    # Instead of comparing against len(entry_df), compare against the sum of 'Sets'
    expected_new_sets = entry_df['Sets'].sum()

    # Validate that the log now includes the new entries with correct values.
    assert len(logger.log_df) == initial_length + expected_new_sets
    assert '2025-02-18' in logger.log_df['date'].values
    assert 'Bench Press' in logger.log_df['exercise_name'].values
    assert logger.log_df['failure'].iloc[0] == True
    assert logger.log_df['log_ID'].max() == initial_max_id + expected_new_sets

def test_reset_entry_sheet(monkeypatch, workout_logger):
    """
    Test that the entry log is reset to an empty DataFrame in memory.
    We assume `reset_entry_sheet()` also clears `logger.entry_df`.
    """
    logger, _, _, _, _ = workout_logger

    # 1) Put some data in entry_df and update the log
    logger.update_log()
    assert not logger.entry_df.empty  # It's not empty now

    # 2) Now reset the entry log
    logger.reset_entry_sheet()

    # 3) Since we've assumed reset_entry_sheet() sets logger.entry_df = empty DataFrame,
    #    we just assert it's empty
    assert logger.entry_df.empty


def test_reset_entry_sheet_preserves_data_validation(tmp_path):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    # 1) Create the Excel file with "entry_sheet"
    entry_file = tmp_path / "entry_with_validation.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "entry_sheet"

    ws["A1"] = "Went to Failure"
    dv = DataValidation(type="list", formula1='"True,False"', allow_blank=True)
    dv.add("A2:A10")
    ws.add_data_validation(dv)

    ws["A2"] = "True"
    ws["A3"] = "False"

    # Create second sheet named "exercise_database"
    wb.create_sheet("exercise_database")
    wb.save(entry_file)

    # 2) Create a minimal JSON file for log_file
    log_file = tmp_path / "mock_log.json"
    log_file.write_text("[]")  # valid JSON array => empty DataFrame
    exercises_file = tmp_path / "mock_exercises.json"
    exercises_file.write_text("[]")  # or some valid JSON content

    # 3) Create the WorkoutLogger with valid paths
    #    (The third argument is also str(entry_file) if you want the same file for exercises)
    logger = WorkoutLogger(str(entry_file), str(log_file), str(exercises_file))

    # 4) load_data() and reset the entry sheet
    logger.load_data()
    logger.reset_entry_sheet()

    # 5) Confirm rows 2+ are cleared but data validation remains
    import openpyxl
    wb_after = openpyxl.load_workbook(entry_file)
    ws_after = wb_after["entry_sheet"]

    assert ws_after["A2"].value is None
    assert ws_after["A3"].value is None

    dv_rules = ws_after.data_validations.dataValidation
    assert len(dv_rules) == 1, "Expected exactly 1 data validation rule"


def test_update_exercise_database_simple(workout_logger):
    """
    Test the simple (single-entry) updating process for the exercise database.
    After loading data and updating the log and exercise database,
    the 'Squat' row should have its 'effectiveness', 'max_weight', and 'max_reps'
    updated, and 'Bench Press' should have an updated 'max_weight'.
    """
    logger, entry_df, new_exercises_df, log_df, exercises_df = workout_logger

    # Update the log and then update the exercise database based on that log.
    logger.update_log()
    logger.update_exercise_database()

    # Retrieve updated values for assertions
    squat = logger.exercises_df.loc[logger.exercises_df['exercise_name'] == 'Squat']
    bench = logger.exercises_df.loc[logger.exercises_df['exercise_name'] == 'Bench Press']

    assert squat['effectiveness'].iloc[0] == 5
    assert squat['max_weight'].iloc[0] == 225
    assert squat['max_reps'].iloc[0] == 10
    assert bench['max_weight'].iloc[0] == 185

def test_update_exercise_database_complex(workout_logger):
    """
    Test the complex (multiple-entry) updating process for the exercise database.
    First the log is updated with the initial mock data, then the entry log is reset.
    A new set of entries is then provided, simulating updated workout data for 'Squat'
    and a new exercise 'Deadlift'. The exercise database should update to the new maximums.
    """
    logger, entry_df, new_exercises_df, log_df, exercises_df = workout_logger

    # Load initial data and update the log (simulate previous entries)
    logger.update_log()
    # Now reset the entry log to simulate a new workout session.
    logger.reset_entry_sheet()

    # Create new entry data with updated values.
    new_entry_data = {
        'Date': ["2025-03-18", "2025-03-19"],
        'Day': ["Tuesday", "Wednesday"],
        'Exercise_ID': [101, 102],
        'Exercise_Name': ["Squat", "Deadlift"],
        'Sets': [4, 5],
        'Reps': [5, 10],
        'Weight': [250, 200],
        'Rest Time (sec)': [90, 120],
        'Effectiveness': [3, 4],
        'Went to failure': [True, False],
        'Notes': ["Felt weak", "Easy"]
    }
    logger.entry_df = pd.DataFrame(new_entry_data)

    # Update log and then update the exercise database
    logger.update_log()
    logger.update_exercise_database()

    # Retrieve updated rows for assertions
    squat = logger.exercises_df.loc[logger.exercises_df['exercise_name'] == 'Squat']
    deadlift = logger.exercises_df.loc[logger.exercises_df['exercise_name'] == 'Deadlift']

    # The new entries should have updated the values (for Squat, new max values are used)
    assert squat['effectiveness'].iloc[0] == 4
    assert squat['max_weight'].iloc[0] == 250
    assert squat['max_reps'].iloc[0] == 5

    # Deadlift should be updated with the new entry values.
    assert deadlift['max_weight'].iloc[0] == 200
    assert deadlift['max_reps'].iloc[0] == 10

def test_save_data(workout_logger):
    """
    Test that save_data writes the correct data to Excel files.
    Instead of writing to disk, we patch pd.ExcelWriter and the DataFrame.to_excel method
    so we can assert that each DataFrame is saved with the correct sheet name.
    """
    logger, entry_df, new_exercises_df, log_df, exercises_df = workout_logger

    # Simulate a session: load, update log and exercise database, and reset the entry log.
    logger.update_log()
    logger.update_exercise_database()
    logger.reset_entry_sheet()

    # Patch ExcelWriter and the to_excel methods to avoid actual file I/O.
    with patch("pandas.ExcelWriter") as mock_writer_class:
        # Create a fake writer instance (simulate context manager)
        fake_writer = mock_writer_class.return_value.__enter__.return_value

        with patch.object(logger.log_df, "to_excel") as mock_log_to_excel, \
             patch.object(logger.exercises_df, "to_excel") as mock_exercises_to_excel:

            logger.save_data()

            # Assert that each DataFrame's to_excel method was called with the fake writer and correct sheet name.
            mock_log_to_excel.assert_called_once_with(fake_writer, sheet_name='workout_log', index=False)
            mock_exercises_to_excel.assert_called_once_with(fake_writer, sheet_name='exercise_database', index=False)
