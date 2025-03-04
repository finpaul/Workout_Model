import pandas as pd
import openpyxl


class WorkoutLogger:
    def __init__(self, entry_file, log_file, exercises_file):
        self.entry_file = entry_file
        self.log_file = log_file
        self.exercises_file = exercises_file

    def load_data(self):
        """Loads the workout entry table and workout log database."""
        self.entry_df = pd.read_excel(self.entry_file, sheet_name='entry_sheet')
        self.new_exercises_df = pd.read_excel(self.entry_file, sheet_name='exercise_database')
        self.log_df = pd.read_json(self.log_file)
        self.exercises_df = pd.read_json(self.exercises_file)

    def update_log(self):
        """
        Appends new workout entries from the entry sheet to the workout log database.
        If `Reps` or `Weight` are comma-separated strings, we expand them into multiple rows.
        """
        new_entries = []

        # Determine starting log_ID
        if self.log_df.empty:
            start_id = 1
        else:
            start_id = int(self.log_df['log_ID'].max() + 1)

        for idx, row in self.entry_df.iterrows():
            # Parse Reps and Weight columns (handle single or multiple sets)
            reps_str = str(row['Reps'])  # e.g. "10,8,6,4"
            weight_str = str(row['Weight'])  # e.g. "225,225,225,250"

            # Split on commas
            reps_list = [r.strip() for r in reps_str.split(',')]
            weight_list = [w.strip() for w in weight_str.split(',')]

            # If the user provided a single number for all sets, we can replicate it
            # e.g., if row['Sets'] = 4 but Weight = "225", replicate "225" 4 times.
            if len(reps_list) == 1 and row['Sets'] > 1:
                reps_list = reps_list * row['Sets']
            if len(weight_list) == 1 and row['Sets'] > 1:
                weight_list = weight_list * row['Sets']

            # Safety check: if sets do not match the length of the lists, you can raise an error or handle it gracefully
            if len(reps_list) != row['Sets'] or len(weight_list) != row['Sets']:
                raise ValueError(
                    f"Number of sets ({row['Sets']}) does not match length of reps/weight for row:\n{row}"
                )

            # Now create one new log entry (row) per set
            for set_index in range(row['Sets']):
                new_entry = {
                    'log_ID': start_id,
                    'exercise_ID': row['Exercise_ID'],
                    'exercise_name': row['Exercise_Name'],
                    'date': row['Date'],
                    'sets': 1,  # each row in the log will represent a single set
                    'reps': float(reps_list[set_index]),
                    'weight': float(weight_list[set_index]),
                    'rest_time': row['Rest Time (sec)'],
                    'effectiveness': row['Effectiveness'],
                    'failure': row['Went to failure'],
                    'notes': row['Notes']
                }
                new_entries.append(new_entry)
                start_id += 1

        # Convert the list of dictionaries into a DataFrame
        new_entries_df = pd.DataFrame(new_entries)

        # Append these new entries to self.log_df
        self.log_df = pd.concat([self.log_df, new_entries_df], ignore_index=True)

        # Save the updated log to JSON (or wherever you normally store it)
        self.log_df.to_json("Workout_Model/src/files/workout_log.json", orient='records', indent=4)

    def reset_entry_sheet(self):
        """Clears out row 2 onward in 'entry_sheet', keeping data validation intact."""
        workbook = openpyxl.load_workbook(self.entry_file)
        ws = workbook['entry_sheet']

        #Completely clear cell values from row 2 downward
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        workbook.save(self.entry_file)

        # ALSO clear the in-memory DataFrame so the test sees it's empty
        self.entry_df = pd.DataFrame(columns=self.entry_df.columns)

        
    def update_exercise_database(self):
        """Adds any new exercises and updates effectiveness and last weight in the Exercises Database from the log."""
        # Ensure that exercise_ID is the index for proper alignment.
        self.new_exercises_df.set_index('exercise_ID', inplace=True)
        self.exercises_df.set_index('exercise_ID', inplace=True)

        # Identify and add any new exercises.
        new_exercises = self.new_exercises_df.loc[~self.new_exercises_df.index.isin(self.exercises_df.index)]
        self.exercises_df = pd.concat([self.exercises_df, new_exercises], axis=0)

        # Update average effectiveness for exercises present in the log.
        avg_effectiveness = self.log_df.groupby('exercise_ID')['effectiveness'].mean()
        self.exercises_df.loc[avg_effectiveness.index, 'effectiveness'] = avg_effectiveness

        # Find the maximum weight (and corresponding reps) for each exercise from the log.
        max_weight_idx = self.log_df.groupby('exercise_ID')['weight'].idxmax()
        max_weight_reps = self.log_df.loc[max_weight_idx, ['exercise_ID', 'weight', 'reps']]
        max_weight_reps.set_index('exercise_ID', inplace=True)
        self.exercises_df.loc[max_weight_reps.index, 'max_weight'] = max_weight_reps['weight']
        self.exercises_df.loc[max_weight_reps.index, 'max_reps'] = max_weight_reps['reps']

        # Now reset the index so that exercise_ID becomes a column.
        self.exercises_df.reset_index(inplace=True)

        # Save the updated exercise database.
        self.exercises_df.to_json("Workout_Model/src/files/exercise_database.json", orient='records', indent=4)

    def save_data(self):
        """Saves updated logs and exercises database to respective sheets in their files."""
        with pd.ExcelWriter(self.entry_file, engine='openpyxl', mode='w') as writer:
            self.log_df.to_excel(writer, sheet_name='workout_log', index=False)

        with pd.ExcelWriter(self.entry_file, engine='openpyxl', mode='w') as writer:
            self.exercises_df.to_excel(writer, sheet_name='exercise_database', index=False)

    def run(self):
        """Runs the full update process."""
        self.load_data()
        self.update_log()
        self.update_exercise_database()
        self.reset_entry_sheet()
        self.save_data()
